"""
Exportadores - CSV, JSON, Excel
"""

from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime
from loguru import logger

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.database import Lead, LeadStatus, LeadPrioridade, LeadTipo


class BaseExporter:
    """Classe base para exportadores."""

    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _get_filename(self, format_ext: str, prefix: str = "leads") -> Path:
        """Gera nome de arquivo com timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.export_dir / f"{prefix}_{timestamp}.{format_ext}"

    def _lead_to_dict(self, lead: Lead) -> dict:
        """Converte Lead para dict plano para exportacao."""
        return {
            "id": lead.id,
            "post_id": lead.post_id,
            "post_url": lead.post_url or "",
            "post_texto": lead.post_texto or "",
            "post_data": lead.post_data.isoformat() if lead.post_data else "",
            "autor_nome": lead.autor_nome or "",
            "autor_id": lead.autor_id or "",
            "autor_url": lead.autor_url or "",
            "nicho": lead.nicho or "",
            "tipo": lead.tipo.value if lead.tipo else "",
            "status": lead.status.value if lead.status else "",
            "prioridade": lead.prioridade.value if lead.prioridade else "",
            "score": lead.score,
            "valor_mencionado": lead.valor_mencionado or "",
            "keywords_encontradas": json.dumps(lead.keywords_encontradas, ensure_ascii=False) if lead.keywords_encontradas else "",
            "palavras_urgencia": json.dumps(lead.palavras_urgencia, ensure_ascii=False) if lead.palavras_urgencia else "",
            "whatsapp": (lead.contato_info or {}).get("whatsapp", ""),
            "email": (lead.contato_info or {}).get("email", ""),
            "telefone": (lead.contato_info or {}).get("telefone", ""),
            "instagram": (lead.contato_info or {}).get("instagram", ""),
            "notas": lead.notas or "",
            "capturado_em": lead.capturado_em.isoformat() if lead.capturado_em else "",
            "atualizado_em": lead.atualizado_em.isoformat() if lead.atualizado_em else "",
        }


class CSVExporter(BaseExporter):
    """Exporta leads para CSV."""

    def export(self, leads: list[Lead], filename: str = None) -> str:
        """Exporta leads para CSV. Retorna caminho do arquivo."""
        filepath = Path(filename) if filename else self._get_filename("csv")
        
        data = [self._lead_to_dict(lead) for lead in leads]
        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding="utf-8-sig")
        
        logger.info(f"Exportados {len(leads)} leads para CSV: {filepath}")
        return str(filepath)


class JSONExporter(BaseExporter):
    """Exporta leads para JSON."""

    def export(self, leads: list[Lead], filename: str = None) -> str:
        """Exporta leads para JSON. Retorna caminho do arquivo."""
        filepath = Path(filename) if filename else self._get_filename("json")
        
        data = [self._lead_to_dict(lead) for lead in leads]
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Exportados {len(leads)} leads para JSON: {filepath}")
        return str(filepath)


class ExcelExporter(BaseExporter):
    """Exporta leads para Excel com formatacao."""

    # Cores por prioridade
    FILL_URGENTE = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    FILL_ALTA = PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid")
    FILL_MEDIA = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    FILL_BAIXA = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def export(self, leads: list[Lead], filename: str = None) -> str:
        """Exporta leads para Excel formatado. Retorna caminho do arquivo."""
        filepath = Path(filename) if filename else self._get_filename("xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Headers
        headers = [
            "ID", "Score", "Prioridade", "Tipo", "Status", "Nicho",
            "Autor", "Texto", "Keywords", "Urgencia",
            "WhatsApp", "Email", "Telefone", "Instagram",
            "Valor", "Data Post", "URL Post", "URL Autor",
            "Notas", "Capturado em",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.BORDER
        
        # Linhas
        for row_idx, lead in enumerate(leads, 2):
            d = self._lead_to_dict(lead)
            contato = lead.contato_info or {}
            
            row_data = [
                lead.id,
                lead.score,
                lead.prioridade.value if lead.prioridade else "",
                lead.tipo.value if lead.tipo else "",
                lead.status.value if lead.status else "",
                lead.nicho or "",
                lead.autor_nome or "",
                lead.post_texto or "",
                d.get("keywords_encontradas", ""),
                d.get("palavras_urgencia", ""),
                contato.get("whatsapp", ""),
                contato.get("email", ""),
                contato.get("telefone", ""),
                contato.get("instagram", ""),
                lead.valor_mencionado or "",
                d.get("post_data", ""),
                lead.post_url or "",
                lead.autor_url or "",
                lead.notas or "",
                d.get("capturado_em", ""),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Cor da linha por prioridade
            fill_map = {
                LeadPrioridade.URGENTE: self.FILL_URGENTE,
                LeadPrioridade.ALTA: self.FILL_ALTA,
                LeadPrioridade.MEDIA: self.FILL_MEDIA,
                LeadPrioridade.BAIXA: self.FILL_BAIXA,
            }
            fill = fill_map.get(lead.prioridade)
            if fill:
                for col in range(1, 3):  # Colore Score + Prioridade
                    ws.cell(row=row_idx, column=col).fill = fill
        
        # Ajusta larguras
        col_widths = [6, 8, 12, 14, 12, 14, 25, 50, 30, 20, 20, 25, 18, 18, 12, 18, 45, 45, 30, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        
        # Freeze panes e autofilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        wb.save(filepath)
        logger.info(f"Exportados {len(leads)} leads para Excel: {filepath}")
        return str(filepath)


# Factory
EXPORTERS = {
    "csv": CSVExporter,
    "json": JSONExporter,
    "xlsx": ExcelExporter,
}


def export_leads(leads: list[Lead], format: str = "xlsx", export_dir: str = "exports", filename: str = None) -> str:
    """
    Exporta leads no formato especificado.
    Retorna caminho do arquivo gerado.
    """
    exporter_cls = EXPORTERS.get(format)
    if not exporter_cls:
        raise ValueError(f"Formato invalido: {format}. Use: {list(EXPORTERS.keys())}")
    
    exporter = exporter_cls(export_dir=export_dir)
    return exporter.export(leads, filename=filename)
